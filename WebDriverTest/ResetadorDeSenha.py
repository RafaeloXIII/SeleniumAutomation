from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime, timedelta
from chrome_manager.chrome_manager import WebDriverManager
import json
import csv
import openpyxl
import pyperclip

# Le arquivo config.json

with open('WebDriverTest\config.json', 'r') as config_file:
    config = json.load(config_file)

# Define your download directory
download_directory = config.get('default_dowload_directory')

# Create a WebDriverManager instance
driver_manager = WebDriverManager(download_directory)

# Create a WebDriver instance for your automation
navegador = driver_manager.create_driver()

# Now you can use the 'navegador' WebDriver instance in your automation script
navegador.get("https://jmsbr.jtjms-br.com/index")
time.sleep(10)

data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# InfoBasicas
indicador_element = navegador.find_element(By.CSS_SELECTOR, '#gateway > div.index.not-activity.dark > div.wrap-main > div > div.box-wrap > div.box-left > div.box-flex > div:nth-child(1)')
indicador_element.click()
time.sleep(3)

# Estrutura org.
estrutura_element = navegador.find_element(By.CSS_SELECTOR, '#gateway > div.home > div.homeLeft.el-scrollbar.mini > div.el-scrollbar__wrap > div > div.sidebar.siderbar-mini > div.sidebar-body > ul.el-menu-vertical-left.el-menu > li:nth-child(5)')
estrutura_element.click()
time.sleep(3)

#Info do Colaborador
monitoramento_movimentacao = navegador.find_element(By.CSS_SELECTOR, '#gateway > div.home > div.homeLeft.el-scrollbar > div.el-scrollbar__wrap > div > div.sidebar > div.sidebar-body > ul.el-menu-vertical-right.el-menu > li:nth-child(1)')
monitoramento_movimentacao.click()              
time.sleep(3)



# # xlsx_file = r"Z:\Recursos\Teste\example.xlsx"  # Provide your xlsx file path here
xlsx_file = r"C:\Users\Rafael Braga\Downloads\Senhas Resetadas(5).xlsx"

workbook = openpyxl.load_workbook(xlsx_file)
sheet = workbook.active

codes = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    codes.append(row[8]) 


code_counter = 0
for code in codes[code_counter:]:

    if code is None or code == "":
        print("No more elements")
        break
    
    print(code)
    #Numero Input
    numero_input = navegador.find_element(By.CSS_SELECTOR, '#jmsSearch > form > div:nth-child(1) > div > div > div > input')
    numero_input.click()

    time.sleep(3)
    pyperclip.copy(code)
    numero_input.click()
    # numero_input.send_keys(Keys.CONTROL, 'v')
    # numero_input.send_keys(Keys.ENTER)
    numero_input.send_keys(code)
    code_counter+=1
    time.sleep(2)

    #Pesquisar
    pequisar_element = navegador.find_element(By.CSS_SELECTOR, '#jmsTopMenu > div.avue-crud__left > button.el-button.btn-query.el-button--primary.el-button--small')
    pequisar_element.click()              
    time.sleep(5)
    try:
        # Reset
        editar_element = navegador.find_element(By.CSS_SELECTOR,'#basicData > div > div.user-list.avue-crud > div.el-table.el-table--fit.el-table--striped.el-table--border.el-table--scrollable-x.el-table--enable-row-transition.el-table--small > div.el-table__fixed-right > div.el-table__fixed-body-wrapper > table > tbody > tr > td.el-table_1_column_16 > div > button:nth-child(3)')
        editar_element.click()
        time.sleep(0.5)
    except:
        # Remover
        limpar_element = navegador.find_element(By.CSS_SELECTOR, '#jmsTopMenu > div.avue-crud__left > button:nth-child(2)')
        limpar_element.click()
        # remover_element = navegador.find_element(By.CSS_SELECTOR, '#jmsSearch > form > div:nth-child(1) > div > div > div > div > span > a')
        # remover_element.click()
        time.sleep(1)   
        continue

    # Confirm JMS
    confirm_element = navegador.find_element(By.CSS_SELECTOR,'#basicData > div > div.el-dialog__wrapper.user-dialog > div > div.el-dialog__body > div > div.dialog-footer > div > button.el-button.confirm.el-button--primary.el-button--small')
    confirm_element.click()
    time.sleep(1)

    # Copy
    copy_element = navegador.find_element(By.CSS_SELECTOR, '#basicData > div > div.el-dialog__wrapper.user-dialog > div > div.el-dialog__body > div > div:nth-child(1) > div.lgd-text.pull-left > span > i')
    copy_element.click()
    time.sleep(0.5)

    copied_code = pyperclip.paste().strip()
    sheet.cell(row=code_counter+1, column=10).value = copied_code
    workbook.save("Senhas Resetadas.xlsx")

    # Confirm
    confirm_element = navegador.find_element(By.CSS_SELECTOR,'#basicData > div > div.el-dialog__wrapper.user-dialog > div > div.el-dialog__body > div > div.dialog-footer > div > button')
    confirm_element.click()
    time.sleep(1)

    # Remover
    limpar_element = navegador.find_element(By.CSS_SELECTOR, '#jmsTopMenu > div.avue-crud__left > button:nth-child(2)')
    limpar_element.click()
    # remover_element = navegador.find_element(By.CSS_SELECTOR, '#jmsSearch > form > div:nth-child(1) > div > div > div > div > span > a')
    # remover_element.click()
    time.sleep(1)

    print(code_counter)

# code_counter = 0

print("Saindo do loop")
time.sleep(30)

navegador.quit()